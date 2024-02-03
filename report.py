# Study Participant Subsidy Reporting Automation (SPSRA)
# Author: Jesse Kaczmarski
# Version: 0.1 (2024-01-23)
# Org: Alaska Center for Energy and Power

# Import required packages
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning) 
import argparse
from datetime import datetime, timedelta
import pandas as pd
import os
import sys
from openpyxl import load_workbook
import subprocess
import shutil

# Argument Parser
parser = argparse.ArgumentParser(prog='Study Participant Subsidy Reporting Automation (SPSRA)',
                                 description='Produces subsidy reports for electric thermal storage heaters used by participants in the ACEP/NSF NNA PM2.5 project.',
                                 epilog='GNU Affero General Public License v3.0')

def valid_date_inputs(a):
    try:
        return datetime.strptime(a, "%Y%m%d")
    except ValueError:
        date_error_message = "Not a valid date input: {0!r}".format(s)
        raise argparse.ArgumentTypeError(date_error_message)

parser.add_argument('startdate',
                    type=valid_date_inputs, 
                    help='ISO 8601 Short Date (e.g., YYYYMMDD)')
parser.add_argument('enddate',
                    type=valid_date_inputs, 
                    help='ISO 8601 Short Date (e.g., YYYYMMDD)')
parser.add_argument('effectiverate',
                    type=float, 
                    help='GVEA Effective Rate $/kWh (e.g., 0.25141)')
parser.add_argument('targetrate',
                    help='Participant Rate $/kWh (e.g., 0.10)',
                    type=float)

args = parser.parse_args()

if args.startdate > args.enddate:
    print('Error: Provided start date exceeds the end date.')
    exit()
else:
    pass

# Status and Interrupt Message
print("Preparing the following subsidy report:",os.linesep,
      "Billing Cycle:",args.startdate.date(), "to", args.enddate.date(), "(", (args.enddate - args.startdate).days + 1,"days )",os.linesep,
      "GVEA Effective Rate:", args.effectiverate,os.linesep,
      "Subsidized Rate:", args.targetrate)
if not input("Would you like to continue? (y/n): ").lower().strip()[:1] == "y": 
    sys.exit(1) # thank you to maroc81: https://gist.github.com/garrettdreyfus/8153571?permalink_comment_id=2614118#gistcomment-2614118

##########################################################################################
### Section 0: Directory Validation
##########################################################################################
try:
    os.makedirs(os.path.join(os.curdir, 'pii', 'reports'))
    print("Sensor data directory has been created at", sensor_dir)
except FileExistsError:
    pass
try:
    os.makedirs(os.path.join(os.curdir, 'pii', 'sensor-data'))
except FileExistsError:
    pass

##########################################################################################
### Section 1: Sensor Data Pull and Store
##########################################################################################

sensor_dir = os.path.join(os.curdir, 'pii', 'sensor-data') # sensor data directory
sensor_url = os.path.join(os.curdir, 'pii','sensor-url.txt') # sensor url to download data

with open(sensor_url) as pii_sensor: call_me = pii_sensor.readlines() # extract BMON sensor URL

print("Downloading sensor data...")
df = pd.read_excel(call_me[0]) # load sensor data into memory

# Cleaning the sensor data
new_cols = df.columns.tolist() # store current columns in list
for i in range(len(new_cols)): new_cols[i] = new_cols[i].replace(", Watts","") # remove extra text in column headings

new_cols[0] = new_cols[0].replace("Timestamp","datetime")
col_rename = {i:j for i,j in zip(df.columns.tolist(),new_cols)} # create dictionary between columns
df.rename(columns=col_rename, inplace=True) # replace columns with clean text version

df = df[df['datetime'] >= args.startdate]
df = df[df['datetime'] <= (args.enddate + timedelta(hours = 23.5))]

sensor_data_output_path = sensor_dir + "sensor-data-" + args.startdate.strftime('%Y%m%d') + "_" + args.enddate.strftime('%Y%m%d') + ".csv"

df.to_csv(sensor_data_output_path, index=False)

print("Sensor data downloaded to", sensor_data_output_path)

##########################################################################################
### Section 2: Calculate subsidies
##########################################################################################
if not input("Would you like to create the subsidy reports with this data? (y/n): ").lower().strip()[:1] == "y": sys.exit(1)

# Filename declarations
subsidy_report_filename = "subsidy-report-" + args.startdate.strftime('%Y%m%d') + "_" + args.enddate.strftime('%Y%m%d') + ".csv"
pr_account_subsidy_filename = "gvea-account-subsidies-" + args.startdate.strftime('%Y%m%d') + "_" + args.enddate.strftime('%Y%m%d') + ".csv"


# Directory checks
report_path = os.path.join(path,'pii','reports', (args.startdate.strftime('%Y%m%d') + "_" + args.enddate.strftime('%Y%m%d') + "/"))
participant_info_path = os.path.join(path,'pii','participant-info.csv')

try:
    os.mkdir(report_path)
    print("Created report folder at", report_path)
except FileExistsError: pass # is this the best way to handle this? examin edge cases

pii = pd.read_csv(participant_info_path) # loading in the participant data
pii['meter_label_2'] = pii['meter_label_2'].fillna('None')

print('Calculating account subsidies')
for i in range(len(pii)):
    # Single Meter Homes
    if pii.loc[i, 'meter_label_2'] == "None":
        pii.loc[i, 'etsh_kwh_usage'] = df[pii.loc[i, 'meter_label_1']].sum() / 1000
    # Dual Meter Homes
    else:
        pii.loc[i,'etsh_kwh_usage'] = (df[pii.loc[i, 'meter_label_1']].sum() + df[pii.loc[i, 'meter_label_2']].sum()) / 1000
    # Usage Based Subsidy Amounts
    pii.loc[i,'etsh_kwh_unsub_cost'] = pii.loc[i,'etsh_kwh_usage'] * args.effectiverate
    pii.loc[i,'etsh_kwh_sub_cost'] = pii.loc[i,'etsh_kwh_usage'] * args.targetrate
    pii.loc[i,'account_credit'] = pii.loc[i,'etsh_kwh_unsub_cost'] - pii.loc[i,'etsh_kwh_sub_cost']

individual_reports_path = os.path.join(report_path, 'individual-reports/')
subsidy_report_output_path = report_path + "subsidy-report-" + args.startdate.strftime('%Y%m%d') + "_" + args.enddate.strftime('%Y%m%d') + ".csv"

# Subsidy data for use in document creation
##########################################################################################
try:
    os.mkdir(individual_reports_path)
except FileExistsError: pass

pii.to_csv(subsidy_report_output_path, index = False)
print('Saved account subsidies to', subsidy_report_output_path)

# Subsidy data for use in purchase request
##########################################################################################
pr_path = os.path.join(report_path, 'purchase-request/')
try:
    os.mkdir(pr_path)
except FileExistsError: pass

pr_pii = pii.loc[:, ('id', 'account_credit')]
pr_pii.rename(columns={'id':'Account Number','account_credit':'Account Credit'}, inplace = True)
pr_pii = pr_pii.set_index('Account Number')
pr_pii['Account Credit'] = pr_pii['Account Credit'].round(2)
pr_pii.loc['Total'] = pr_pii['Account Credit'].sum()
pr_pii.to_csv(pr_path + 'gvea-account-subsidies.csv')
pr_path_saved = pr_path + 'gvea-account-subsidies.csv'
print('Saved GVEA account subsidies to', pr_path_saved)

##########################################################################################
### Section 3: Autofill the purchase request form
##########################################################################################
pr_template_path = os.path.join(path, 'pii','G14304-PR-GVEA-YYYYMMDD.xlsx')

pr_wb = load_workbook(pr_template_path)

ws = pr_wb.active
ws['N10'] = pii['account_credit'].sum().round(2) # Import total PR request amount
ws['N34'] = pii['account_credit'].sum().round(2) # Import total PR request amount
ws['N20'] = pii['account_credit'].sum().round(2) # Import total PR request amount
ws['K20'] = pii['etsh_kwh_usage'].sum().round(2) # Import total kWh being subsidized
ws['M20'] = args.effectiverate - args.targetrate # Import subsidy amount per kWh
ws['H37'] = datetime.now().date().strftime("%Y-%m-%d") # Add the current date to the document

pr_wb_path = pr_path + 'G14304-PR-GVEA-' + datetime.now().date().strftime("%Y%m%d") + '.xlsx'
pr_wb.save(pr_wb_path)
print('Saved purchase request to', pr_wb_path)

##########################################################################################
### Section 4: Create reports for each participant
##########################################################################################

# Create tex files in a temporary folder
latex_template_path = os.path.join(os.curdir,'misc','main.tex')
difference = args.effectiverate - args.targetrate

try:
    os.mkdir(individual_reports_path)
except FileExistsError:
    files = os.listdir(individual_reports_path)
    for f in files:
        os.remove(individual_reports_path + f)

for i in range(len(pii)):
    filename = individual_reports_path + str(pii.loc[i,'filename']) + '-report-' + args.startdate.date().strftime("%Y%m%d") + '_' + args.enddate.date().strftime("%Y%m%d") + '.tex'
    with open(latex_template_path,'r') as report:
        text = report.read()
        text = text.replace('startdate', args.startdate.date().strftime("%b %d, %Y"))
        text = text.replace('enddate', args.enddate.date().strftime("%b %d, %Y"))
        text = text.replace('name', pii.loc[i,'name'])
        text = text.replace('address1', pii.loc[i,'address_1'])
        text = text.replace('address2', pii.loc[i,'address_2'])
        text = text.replace('email', pii.loc[i,'email'])
        text = text.replace('phone', pii.loc[i,'phone'])
        text = text.replace('effectiverate', str(args.effectiverate))
        text = text.replace('targetrate', format(args.targetrate, '.2f'))
        text = text.replace('difference', str(round(difference, 5)))
        text = text.replace('etshkwhusage', format(pii.loc[i,'etsh_kwh_usage'], '.2f'))
        text = text.replace('etshkwhunsubcost', format(pii.loc[i,'etsh_kwh_unsub_cost'], '.2f'))
        text = text.replace('etshkwhsubcost', format(pii.loc[i,'etsh_kwh_sub_cost'], '.2f'))
        text = text.replace('accountcredit', format(pii.loc[i,'account_credit'], '.2f'))

    with open(filename,'w') as output:
        output.write(text)

# Run all tex files through pdflatex
run_me = os.listdir(individual_reports_path)
shutil.copytree('./misc/figures/', individual_reports_path, dirs_exist_ok = True)
original_dir = os.getcwd()
os.chdir(individual_reports_path)
for f in run_me:
    subprocess.run(['pdflatex',f],stdout=subprocess.DEVNULL,stderr=subprocess.STDOUT)

print('Created individual subsidy reports in', individual_reports_path)
os.chdir(original_dir)

# Delete non-pdf files
updated_dir = os.listdir(individual_reports_path)
updated_dir_keep = [f for f in updated_dir if (f.endswith(".pdf"))]
for f in updated_dir:
    if f not in updated_dir_keep:
        os.remove(individual_reports_path + f)
    else:
        pass

exit()
